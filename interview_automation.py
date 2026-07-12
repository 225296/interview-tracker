#!/usr/bin/env python3
"""
Interview tracker automation — engine.

Pipeline applied to an imported candidate file:

  1. Keep only rows whose Interview Status is one of the 6 valid values.
  2. For 'Scheduled skill' / 'Scheduled final' rows, only those dated TODAY
     stay "in scope"; other statuses ignore the date.
  3. Primary skills (SAP, Oracle, Salesforce, Agentforce) -> distributed
     equally across the 6 HR.
  4. All other skills -> also distributed equally across the 6 HR.
  5. POC column = the assigned HR name (only for in-scope rows).
  6. Updates column = dropdown with the 9 disposition values.

Rows that fall out of scope are kept on the sheet with POC left blank
(and lightly greyed), so nothing is lost.

CLI:
    python3 interview_automation.py --template          # blank template
    python3 interview_automation.py <input.xlsx>        # process a file
"""

import sys
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------- #
# Configuration — edit these to change dropdowns, HR count, or skills.
# --------------------------------------------------------------------------- #

PANELS = [f"HR {i}" for i in range(1, 7)]           # fallback if none supplied

PRIMARY_SKILLS = ["SAP", "Oracle", "Salesforce", "Agentforce"]

INTERVIEW_STATUS = [
    "Scheduled skill interview",
    "Scheduled final interview",
    "Final selected",
    "Skill selected",
    "Skill status pending",
    "Final status pending",
]

UPDATES = [
    "Reject",
    "Skill mismatch",
    "OHCM",
    "PHC",
    "Moved to final",
    "Moved to skill 2",
    "Cand renege",
    "EI renege",
    "Tech issue",
]

VALID_STATUSES = {s.lower() for s in INTERVIEW_STATUS}
SCHEDULED_STATUSES = {"scheduled skill interview", "scheduled final interview"}

COLUMNS = [
    "Candidate ID",
    "Candidate Name",
    "Primary Skill",
    "Interview Status",
    "Interview Date",
    "POC",
    "Updates",
]

# Flexible header matching for imported files.
ALIASES = {
    "candidate id": "Candidate ID", "id": "Candidate ID", "emp id": "Candidate ID",
    "candidate name": "Candidate Name", "name": "Candidate Name",
    "candidate": "Candidate Name",
    "primary skill": "Primary Skill", "skill": "Primary Skill", "skills": "Primary Skill",
    "interview status": "Interview Status", "status": "Interview Status",
    "interview date": "Interview Date", "date": "Interview Date",
    "poc": "POC", "point of contact": "POC",
    "updates": "Updates", "update": "Updates", "disposition": "Updates",
    "update / disposition": "Updates",
}

# --------------------------------------------------------------------------- #
# Styling
# --------------------------------------------------------------------------- #

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
POC_FILL = PatternFill("solid", fgColor="E2EFDA")       # in-scope / assigned
OUT_FILL = PatternFill("solid", fgColor="F2F2F2")       # out-of-scope (greyed)
OUT_FONT = Font(color="9AA5B1")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _norm(text):
    return str(text).strip().lower() if text is not None else ""


def as_date(value):
    """Coerce a cell value to a date, or None."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y",
                    "%d-%b-%Y", "%d %b %Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def is_primary(skill):
    return _norm(skill) in {_norm(p) for p in PRIMARY_SKILLS}


# --------------------------------------------------------------------------- #
# Scope + distribution (steps 1-5)
# --------------------------------------------------------------------------- #

def in_scope(row, today):
    """Steps 1-2: valid status, and scheduled rows must be dated today."""
    status = _norm(row.get("Interview Status"))
    if status not in VALID_STATUSES:
        return False
    if status in SCHEDULED_STATUSES:
        return as_date(row.get("Interview Date")) == today
    return True


def assign_pocs(rows, panels=PANELS, today=None):
    """Assign POC (HR) to in-scope rows, equally across panels.

    Primary skills are distributed first (steps 3), then the rest (step 4),
    with one continuous cursor so the overall load stays even (n/6 each).
    Returns the number of in-scope rows.
    """
    today = today or date.today()
    scope = [r for r in rows if in_scope(r, today)]

    groups = defaultdict(list)
    for r in scope:
        groups[r.get("Primary Skill") or "Unspecified"].append(r)

    primary = [s for s in PRIMARY_SKILLS if s in groups]
    primary += [s for s in groups if is_primary(s) and s not in primary]
    rest = sorted(s for s in groups if not is_primary(s))
    ordered = primary + rest

    cursor = 0
    for skill in ordered:
        for r in groups[skill]:
            if not r.get("POC"):                          # respect manual picks
                r["POC"] = panels[cursor % len(panels)]
            cursor += 1
    return len(scope)


# --------------------------------------------------------------------------- #
# Dropdowns (steps 5-6): hidden Lists sheet + validations
# --------------------------------------------------------------------------- #

def _write_lists_sheet(wb, panels=PANELS):
    ws = wb.create_sheet("Lists")
    ws.sheet_state = "hidden"
    named = {
        "A": ("SkillList", PRIMARY_SKILLS),
        "B": ("StatusList", INTERVIEW_STATUS),
        "C": ("PanelList", panels),
        "D": ("UpdatesList", UPDATES),
    }
    ranges = {}
    for col, (title, values) in named.items():
        ws[f"{col}1"] = title
        for i, v in enumerate(values, start=2):
            ws[f"{col}{i}"] = v
        ranges[title] = f"Lists!${col}$2:${col}${len(values) + 1}"
    return ranges


def _apply_validations(ws, ranges, n_rows):
    last = n_rows + 1
    letter = {name: get_column_letter(i + 1) for i, name in enumerate(COLUMNS)}
    dropdowns = {
        "Primary Skill": ranges["SkillList"],
        "Interview Status": ranges["StatusList"],
        "POC": ranges["PanelList"],
        "Updates": ranges["UpdatesList"],
    }
    for column, formula in dropdowns.items():
        dv = DataValidation(type="list", formula1=f"={formula}", allow_blank=True)
        dv.error = "Pick a value from the dropdown list."
        dv.errorTitle = "Invalid entry"
        dv.prompt = f"Choose a {column}"
        col = letter[column]
        dv.add(f"{col}2:{col}{last}")
        ws.add_data_validation(dv)


# --------------------------------------------------------------------------- #
# Reading an imported workbook
# --------------------------------------------------------------------------- #

def read_input(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    lookup = {}
    for i, h in enumerate(headers):
        canon = None
        for c in COLUMNS:
            if _norm(h) == _norm(c):
                canon = c
                break
        if canon is None:
            canon = ALIASES.get(_norm(h))
        if canon and canon not in lookup:
            lookup[canon] = i

    rows = []
    for excel_row in ws.iter_rows(min_row=2):
        if all(c.value in (None, "") for c in excel_row):
            continue
        rows.append({c: (excel_row[lookup[c]].value if c in lookup else None)
                     for c in COLUMNS})
    return rows


# --------------------------------------------------------------------------- #
# Writing the tracker
# --------------------------------------------------------------------------- #

def write_tracker(rows, out_path, panels=PANELS, today=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Interview Tracker"
    ranges = _write_lists_sheet(wb, panels)

    for i, name in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    for r_i, record in enumerate(rows, start=2):
        scoped = today is not None and in_scope(record, today)
        for c_i, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=record.get(name))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            if name == "Interview Date" and record.get(name):
                cell.number_format = "dd-mmm-yyyy"
            if name == "POC" and record.get(name):
                cell.fill = POC_FILL
            elif today is not None and not scoped:
                cell.fill = OUT_FILL
                cell.font = OUT_FONT

    _apply_validations(ws, ranges, len(rows))

    widths = [12, 22, 16, 26, 16, 12, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    wb.save(out_path)


def build_template(out_path, n_blank=50, panels=PANELS):
    rows = [{c: None for c in COLUMNS} for _ in range(n_blank)]
    write_tracker(rows, out_path, panels)          # today=None -> no greying


# --------------------------------------------------------------------------- #
# Summary
# --------------------------------------------------------------------------- #

def summarize(rows, panels, today, scope_count):
    by_poc = defaultdict(lambda: defaultdict(int))
    for r in rows:
        if r.get("POC"):
            by_poc[r["POC"]][r.get("Primary Skill") or "?"] += 1

    parked = len(rows) - scope_count
    lines = [f"Today: {today.isoformat()}",
             f"Rows in file: {len(rows)}   In scope (got POC): {scope_count}"
             f"   Parked (POC blank): {parked}",
             "",
             "POC distribution (HR -> counts by skill):"]
    for p in panels:
        skills = by_poc.get(p, {})
        total = sum(skills.values())
        detail = ", ".join(f"{k}:{v}" for k, v in sorted(skills.items()))
        lines.append(f"  {p:<12} total {total:>2}   {detail}")
    return "\n".join(lines)


def compute_stats(rows, panels, today, scope_count):
    """Structured version of the summary for rich UIs."""
    by_poc = defaultdict(lambda: defaultdict(int))
    present = set()
    for r in rows:
        if r.get("POC"):
            sk = r.get("Primary Skill") or "Unspecified"
            by_poc[r["POC"]][sk] += 1
            present.add(sk)

    ordered_skills = ([s for s in PRIMARY_SKILLS if s in present]
                      + sorted(s for s in present if not is_primary(s)))
    distribution = [{"hr": p,
                     "total": sum(by_poc.get(p, {}).values()),
                     "by_skill": dict(by_poc.get(p, {}))}
                    for p in panels]
    return {
        "today": today,
        "n_rows": len(rows),
        "in_scope": scope_count,
        "parked": len(rows) - scope_count,
        "skills": ordered_skills,
        "distribution": distribution,
        "text": summarize(rows, panels, today, scope_count),
    }


def process_file(in_path, out_path, panels=PANELS, today=None):
    today = today or date.today()
    rows = read_input(in_path)
    scope_count = assign_pocs(rows, panels, today)
    write_tracker(rows, out_path, panels, today)
    return len(rows), compute_stats(rows, panels, today, scope_count)


# --------------------------------------------------------------------------- #

def main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 1
    if argv[1] == "--template":
        build_template("Interview_Template.xlsx")
        print("Template written -> Interview_Template.xlsx")
        return 0
    out = "Interview_Tracker_Output.xlsx"
    count, stats = process_file(argv[1], out)
    print(f"Read {count} rows from {argv[1]}")
    print(f"Tracker written -> {out}\n")
    print(stats["text"])
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
